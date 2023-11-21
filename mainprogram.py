import tkinter as tk
from tkinter import ttk
from tkinter import font
import tkinter.messagebox

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os.path


file_exists = os.path.exists('Database.xlsx')

if file_exists == False:
    wb = Workbook()
    wb.save("Database.xlsx")
    wb = load_workbook("Database.xlsx")
    ws = wb.active
    ws['A1'].value = "Order ID"
    ws['B1'].value = "Item"
    ws['C1'].value = "Customer Name"
    ws['D1'].value = "Address"
    ws['E1'].value = "Status"
    wb.save("Database.xlsx")

else:
    print("File present")

table_data = []

wb = load_workbook("Database.xlsx")
ws = wb.active
s = ""
for row in ws.iter_rows(min_row=2, max_col=5):
    r = []
    for cell in row:
        r.append(cell.value)
    table_data.append(r)

print(table_data)

np, nc, nd, ns, ncn = 0, 0, 0, 0, 0

def chv():
    global np
    global nc
    global nd
    global ns
    global ncn

    for i in table_data:
        if i[4] == "Delivered":
            nd+=1
        if i[4] == "Packaging":
            np+=1
        if i[4] == "Confirmed":
            nc+=1
        if i[4] == "Cancelled":
            ncn+=1
        if i[4] == "Shipping":
            ns+=1
chv()

root = tk.Tk()
root.title("SIMPLY - Order Management System")
root.geometry("1600x900")
root.resizable(0, 0)
icon = tk.PhotoImage(file='./Images/logo.png')
root.iconphoto(True, icon)

def delframe():
    for frame in main_frame.winfo_children():
        frame.destroy()


#Dashboard Page

def dashboard():
    delframe()
    chv()
    df = tk.Frame(main_frame, background="#15114a")

    tm = ImageTk.PhotoImage(Image.open('./Images/m1.png').resize((300,40)))
    tl = tk.Label(df, image=tm,text="", anchor="w", background="#15114a")
    tl.image = tm
    tl.pack(pady=20)

    tb = tk.Frame(df, bg="blue", width=1000, height=600)
    tb.pack(side=tk.BOTTOM, anchor='s', pady=10)

    sb1 = tk.Frame(df, bg="#15114a",width=500, height=500)
    sb1.pack(side=tk.LEFT, pady=20)

    sb2 = tk.Frame(df, bg="#15114a",width=300, height=500)
    sb2.pack(side=tk.RIGHT, pady=20)

    fig1, ax = plt.subplots()
    info = ['Confirmed', 'Packaging', 'Shipping', 'Delivered', 'Cancelled']
    order = [nc, np, ns, nd, ncn]
    bar_labels = ['Confirmed', 'Packaging', 'Shipping', 'Delivered', 'Cancelled']
    bar_colors = ['tab:red', 'tab:blue', 'tab:purple', 'tab:orange', 'tab:green']
    ax.bar(info, order, label=bar_labels, color=bar_colors)
    ax.set_ylabel('Number of Orders')
    ax.set_title('Bar Chart')
    canvas1 = FigureCanvasTkAgg(fig1, master = sb1)
    canvas1.get_tk_widget().pack()
    canvas1.draw()

    labels = 'Confirmed', 'Packaging', 'Delivered', 'Shipping', 'Cancelled'
    sizes = [nc, np, nd, ns, ncn]
    fig2, ax = plt.subplots()
    ax.pie(sizes, labels=labels, autopct='%1.1f%%')
    ax.set_title('Pie Chart')
    canvas2 = FigureCanvasTkAgg(fig2, master=sb2)
    canvas2.get_tk_widget().pack()
    canvas2.draw()

    lf = font.Font(family="Times", size=17, weight="bold")
    style = ttk.Style()
    style.configure("Treeview",background="white",foreground="black",relief = 'flat',borderwidth = 0, bd=5, font=("Arial",13),rowheight=30)
    style.configure("Treeview.Heading", background="white", foreground="black", relief='flat', borderwidth=0, bd=5,
                    font=lf)
    style.map("Treeview", background=[('selected',"blue")])

    tree = ttk.Treeview(tb,column=("Order ID", "Item", "Customer", "Address","Status"), show='headings', height=11)

    tree.tag_configure("oddrow",background="white")
    tree.tag_configure("evenrow",background="#b5fdff")

    tree.column("# 1", anchor=tk.CENTER, width=140)
    tree.heading("# 1", text="Order ID")
    tree.column("# 2", anchor=tk.CENTER, width=250)
    tree.heading("# 2", text="Item")
    tree.column("# 3", anchor=tk.CENTER, width=250)
    tree.heading("# 3", text="Customer Name")
    tree.column("# 4", anchor=tk.CENTER, width=500)
    tree.heading("# 4", text="Address")
    tree.column("# 5", anchor=tk.CENTER, width=140)
    tree.heading("# 5", text="Status")
    count = 0
    for i in table_data:
        if count % 2 == 0:
            tree.insert('', 'end', text="1", values=i, tags="evenrow")
        else:
            tree.insert('', 'end', text="1", values=i, tags="oddrow")
        count+=1
    tree.pack()

    df.pack()

#Orders Page

def orders():

    delframe()

    ctd = table_data

    def display_selected(choice):
        if choice != "All Orders":
            global ctd
            choice = clicked.get()
            ctd = []
            for i in table_data:
                if i[4] == choice:
                    ctd.append(i)
            tree.update()

            for item in tree.get_children():
                tree.delete(item)
            c = 0
            for j in ctd:
                if c % 2 == 0:
                    tree.insert('', 'end', text="1", values=j, tags="evenrow")
                else:
                    tree.insert('', 'end', text="1", values=j, tags="oddrow")
                c += 1

            tree.configure(height=c)
        else:
            for item in tree.get_children():
                tree.delete(item)
            c = 0
            for j in table_data:
                if c % 2 == 0:
                    tree.insert('', 'end', text="1", values=j, tags="evenrow")
                else:
                    tree.insert('', 'end', text="1", values=j, tags="oddrow")
                c += 1
            tree.configure(height=c)

    def on_entry_change(*args):
        new_value = entry_var.get()
        ndt = []
        for i in ctd:
            if new_value in i[0]:
                ndt.append(i)

        for item in tree.get_children():
            tree.delete(item)
        c = 0
        for j in ndt:
            if c % 2 == 0:
                tree.insert('', 'end', text="1", values=j, tags="evenrow")
            else:
                tree.insert('', 'end', text="1", values=j, tags="oddrow")
            c += 1
        tree.configure(height=c)

        if ndt == []:
            tkinter.messagebox.showinfo("SIMPLY ALERT", "No Order Found !!")
            entry_var.set("")
            clicked.set("All Orders")
            for item in tree.get_children():
                tree.delete(item)
            c = 0
            for j in table_data:
                if c % 2 == 0:
                    tree.insert('', 'end', text="1", values=j, tags="evenrow")
                else:
                    tree.insert('', 'end', text="1", values=j, tags="oddrow")
                c += 1
            tree.configure(height=c)

    of = tk.Frame(main_frame,background="#15114a")

    orm = ImageTk.PhotoImage(Image.open('./Images/m2.png').resize((250,50)))
    ol = tk.Label(of, image=orm, text="", anchor="w", background="#15114a")
    ol.image = orm

    to = tk.Frame(of, bg="#15114a", width=1000, height=600)
    to.pack(side=tk.BOTTOM, anchor='s', pady=10)

    src = tk.Frame(of, bg="#15114a", width=500, height=200)
    src.pack(side=tk.BOTTOM, anchor='c',pady=10,padx=5)

    sf = font.Font(family="Times", size=20, weight="bold")
    so = tk.Label(src, text="Enter Order ID - ", font=sf, foreground="#fff",background="#15114a", anchor="w")
    so.pack(side=tk.LEFT)

    entry_var = tk.StringVar()
    entry_var.trace("w", on_entry_change)

    search = tk.Entry(src, width=10, font=('calibre',17,'normal'), justify=tk.CENTER, borderwidth=5, border=5, textvariable=entry_var)
    search.pack(side=tk.LEFT,padx=(0,650))

    options = [
        "All Orders",
        "Confirmed",
        "Delivered",
        "Shipping",
        "Packaging",
        "Cancelled"
    ]

    clicked = tk.StringVar()
    clicked.set("All Orders")

    co = tk.Label(src, text="Sort By - ",font=sf, foreground="#fff", background="#15114a", anchor="w")
    co.pack(side=tk.LEFT)

    drop = tk.OptionMenu(src, clicked, *options, command = display_selected)
    drop.config(width=10,height=1,font=of)
    drop.pack(side=tk.RIGHT)

    lf = font.Font(family="Times", size=20, weight="bold")
    style = ttk.Style()
    style.configure("Treeview", background="white", foreground="black", relief='flat', borderwidth=0, bd=5,
                    font=("Arial", 15),rowheight=30)
    style.configure("Treeview.Heading", background="white", foreground="black", relief='flat', borderwidth=0, bd=5,
                    font=lf,rowheight=20)
    style.map("Treeview", background=[('selected', "blue")])

    tree = ttk.Treeview(to, column=("Order ID", "Item", "Customer", "Address", "Status"), show='headings', height=21)

    tree.tag_configure("oddrow", background="white")
    tree.tag_configure("evenrow", background="#b5fdff")

    tree.column("# 1", anchor=tk.CENTER, width=140)
    tree.heading("# 1", text="Order ID")
    tree.column("# 2", anchor=tk.CENTER, width=250)
    tree.heading("# 2", text="Item")
    tree.column("# 3", anchor=tk.CENTER, width=250)
    tree.heading("# 3", text="Customer Name")
    tree.column("# 4", anchor=tk.CENTER, width=500)
    tree.heading("# 4", text="Address")
    tree.column("# 5", anchor=tk.CENTER, width=140)
    tree.heading("# 5", text="Status")
    count = 0
    for i in ctd:
        if count % 2 == 0:
            tree.insert('', 'end', text="1", values=i, tags="evenrow")
        else:
            tree.insert('', 'end', text="1", values=i, tags="oddrow")
        count += 1
    tree.configure(height=count)
    tree.pack()
    ol.pack(pady=(40,30))
    of.pack(pady=20)
    of.mainloop()

#Add Order Page

def addorders():
    delframe()

    def adddata():
        n = namee.get()
        i = iteme.get()
        a = addresse.get()
        o = order.get()
        c = clicked.get()

        if n != "" and i != "" and a != "" and o != "":
            print("Valid")
            table_data.append([o, i, n, a, c])
            ws.append([o, i, n, a, c])
            wb.save("Database.xlsx")
            chv()
            print("Done")
            tkinter.messagebox.showinfo("SIMPLY POPUP", "Added Order to Database Successfully !!")
            namee.delete(0, tk.END)
            iteme.delete(0, tk.END)
            addresse.delete(0, tk.END)
            order.delete(0, tk.END)

        else:
            print("Invalid")
            tkinter.messagebox.showinfo("SIMPLY ERROR", "Please Fill out all the Entry Fields !!")

    adf = tk.Frame(main_frame,background="#15114a")

    adm = ImageTk.PhotoImage(Image.open('./Images/m3.png').resize((330, 45)))
    adl = tk.Label(adf, image=adm,text="", background="#15114a", anchor="n")
    adl.image = adm

    submit = tk.Button(adf, text="SUBMIT", font=font.Font(family="Arial", size=20, weight="bold"),foreground="white",
                       background="#0094ff", width= 20, height=1,command= lambda : adddata())
    submit.pack(side=tk.BOTTOM, pady=(45,0))

    br = tk.Frame(adf, background="#fff",highlightbackground="yellow", highlightthickness=5)

    r3 = tk.Frame(br, bg="#fff", width=500, height=200)
    r3.pack(side=tk.BOTTOM, anchor='c', pady=30, padx=5)

    r1 = tk.Frame(br, bg="#fff", width=500, height=200)
    r1.pack(side=tk.BOTTOM, anchor='c', pady=30, padx=5)

    sf = font.Font(family="Times", size=24, weight="normal")
    so = tk.Label(r1, text="Enter Order ID - ", font=sf,background="#fff", anchor="w")
    so.pack(side=tk.LEFT)

    order = tk.Entry(r1, width=18, font=('calibre', 20, 'normal'), justify=tk.CENTER, borderwidth=5, border=5)
    order.pack(side=tk.LEFT, padx=(0, 250))

    options = [
        "Confirmed",
        "Delivered",
        "Shipping",
        "Packaging",
        "Cancelled"
    ]

    clicked = tk.StringVar()
    clicked.set("Confirmed")

    co = tk.Label(r1, text="Status of Package - ", font=sf,background="#fff", anchor="w")
    co.pack(side=tk.LEFT)

    drop = tk.OptionMenu(r1, clicked, *options)
    drop.config(width=10, height=1, font=font.Font(family="Arial", size=20, weight="normal"))
    drop.pack(side=tk.RIGHT)

    r2 = tk.Frame(br, bg="#fff", width=500, height=200)
    r2.pack(side=tk.BOTTOM, anchor='c', pady=30, padx=5)

    name = tk.Label(r2, text="Enter Customer Name - ", font=sf, background="#fff", anchor="w")
    name.pack(side=tk.LEFT)

    namee = tk.Entry(r2, width=20, font=('calibre', 20, 'normal'), justify=tk.CENTER, borderwidth=5, border=5)
    namee.pack(side=tk.LEFT, padx=(0, 50))

    item = tk.Label(r2, text="Enter Item Name - ", font=sf, background="#fff", anchor="w")
    item.pack(side=tk.LEFT)

    iteme = tk.Entry(r2, width=20, font=('calibre', 20, 'normal'), justify=tk.CENTER, borderwidth=5, border=5)
    iteme.pack(side=tk.LEFT)

    address = tk.Label(r3, text="Enter Address - ", font=sf, background="#fff", anchor="w")
    address.pack(side=tk.LEFT)
    addresse = tk.Entry(r3, width=65, font=('calibre', 20, 'normal'), justify=tk.CENTER, borderwidth=5, border=5)
    addresse.pack()

    br.pack(pady=20, side=tk.BOTTOM)
    adl.pack(pady=(50,60))
    adf.pack(pady=20)

def modorder():
    delframe()
    itin = 0

    def display_selected(choice):
        global itin
        if oid.get() != "":
            print(choice)
            table_data[itin][4] = choice
            print(table_data[itin])
            cin = itin + 2
            cel = "E" + str(cin)
            print(cel)
            ws[cel] = choice
            wb.save("Database.xlsx")
            itin = 0
            tkinter.messagebox.showinfo("SIMPLY POPUP", "Changed Status of Package Successfully !!")
            oid.set("")
            add.set("")
            cus.set("")
            it.set("")
            clicked.set("")
            entry_var.set("")
            chv()

    def search():
        global itin
        sorder = entry_var.get()
        f = False
        ind = 0
        for i in table_data:
            if i[0] == sorder:
                f = True
                ind = table_data.index(i)
                itin = ind
                break

        v = table_data[ind]

        if f:
            print("Found at", ind)

            oid.set("𝐎𝐫𝐝𝐞𝐫 𝐈𝐃 -  "+v[0])
            it.set("𝐈𝐭𝐞𝐦 -  "+v[1])
            cus.set("𝐂𝐮𝐬𝐭𝐨𝐦𝐞𝐫 𝐍𝐚𝐦𝐞 -  "+v[2])
            add.set("𝐀𝐝𝐝𝐫𝐞𝐬𝐬 -  "+v[3])
            clicked.set(v[4])

        else:
            tkinter.messagebox.showinfo("SIMPLY ALERT", "No Order Found !!")
            oid.set("")
            add.set("")
            cus.set("")
            it.set("")
            clicked.set("")
            entry_var.set("")


    def deletedata():
        global itin
        if oid.get() != "":
            print("Delete")
            table_data.pop(itin)
            din = itin + 2
            ws.delete_rows(din)
            wb.save("Database.xlsx")
            itin = 0
            tkinter.messagebox.showinfo("SIMPLY ALERT", "Deleted Order Successfully !!")
            oid.set("")
            add.set("")
            cus.set("")
            it.set("")
            clicked.set("")
            entry_var.set("")
            chv()


    modf = tk.Frame(main_frame,background="#15114a")

    modm = ImageTk.PhotoImage(Image.open('./Images/m4.png').resize((300,40)))
    modl = tk.Label(modf, image=modm,text="", background="#15114a", anchor="w")
    modm.image = modm
    modl.pack(side=tk.TOP)

    entry_var = tk.StringVar()

    sf = tk.Frame(modf, bg="#15114a", width=500, height=200)
    sf.pack(side=tk.TOP,anchor='c', pady=30, padx=5)

    so = tk.Label(sf, text="Enter Order ID - ", font=font.Font(family="Times", size=25, weight="bold"), foreground="#fff",background="#15114a", anchor="w")
    so.pack(side=tk.LEFT)
    soe = tk.Entry(sf, width=20, font=('calibre', 25, 'normal'), justify=tk.CENTER, borderwidth=5, border=5, textvariable=entry_var)
    soe.pack(side=tk.LEFT)
    sob = tk.Button(sf, text = "🔍",width=5,height=1,background="#0094ff",font=('calibre', 20, 'normal'),borderwidth=5, border=5, command=search)
    sob.pack(side=tk.RIGHT, padx=60)

    delete = tk.Button(modf, text="🚫 Delete Order 🚫", font=font.Font(family="Arial", size=20, weight="bold"), foreground="white",
                       background="red", width=20, command=lambda: deletedata())
    delete.pack(side=tk.BOTTOM, pady=(45, 0))

    options = [
        "Confirmed",
        "Delivered",
        "Shipping",
        "Packaging",
        "Cancelled"
    ]

    clicked = tk.StringVar()

    df = tk.Frame(modf, bg="#15114a", width=500, height=200)
    df.pack(side=tk.BOTTOM, anchor='c', pady=30, padx=5)

    dropl = tk.Label(df, text="𝑺𝒕𝒂𝒕𝒖𝒔 𝑶𝒇 𝑷𝒂𝒄𝒌𝒂𝒈𝒆 - ", foreground="#fff",font=font.Font(family="Times", size=25, weight="bold"),
                  background="#15114a", anchor="w")
    dropl.pack(side=tk.LEFT)
    drop = tk.OptionMenu(df, clicked, *options, command=display_selected)
    drop.config(width=10, height=1, font=font.Font(family="Arial", size=20, weight="normal"))
    drop.pack(side=tk.LEFT)

    add = tk.StringVar()
    address = tk.Message(modf, textvariable=add, font=font.Font(family="Arial", size=25, weight="normal"), width=1200,foreground="#fff",
                         background="#15114a", anchor="w", justify=tk.LEFT)
    address.pack(side=tk.BOTTOM, pady=10)

    cus = tk.StringVar()
    cusname = tk.Label(modf, textvariable=cus, font=font.Font(family="Arial", size=25, weight="normal"),foreground="#fff",
                         background="#15114a", anchor="w")
    cusname.pack(side=tk.BOTTOM, pady=10)

    it = tk.StringVar()
    itm = tk.Label(modf, textvariable=it, font=font.Font(family="Arial", size=25, weight="normal"),foreground="#fff",
                         background="#15114a", anchor="w")
    itm.pack(side=tk.BOTTOM, pady=10)

    oid = tk.StringVar()
    oidl = tk.Label(modf, textvariable=oid, font=font.Font(family="Arial", size=25, weight="normal"),foreground="#fff",
                     background="#15114a", anchor="w")
    oidl.pack(side=tk.BOTTOM, pady=10)

    modl.pack(pady=40)
    modf.pack(pady=20)

def abt():
    delframe()
    abtf = tk.Frame(main_frame,background="#003459", highlightbackground="#af6bef", highlightthickness=4)

    mfr = tk.Frame(abtf, background="#003459")

    m1 = ImageTk.PhotoImage(Image.open('./Images/logo.png').resize((300, 300)))
    m1l = tk.Label(mfr, image=m1, text="", background="#003459", anchor="w")
    m1l.image = m1
    m1l.pack(side=tk.LEFT, padx=(40,0), pady=20)

    m2 = ImageTk.PhotoImage(Image.open('./Images/logoname.png').resize((350, 55)))
    m2l = tk.Label(mfr, image=m2, text="", background="#003459", anchor="w")
    m2l.image = m2
    m2l.pack(side=tk.LEFT, padx=(50,0))

    m3 = ImageTk.PhotoImage(Image.open('./Images/tg2.png').resize((550, 40)))
    m3l = tk.Label(mfr, image=m3, text="", background="#003459", anchor="w")
    m3l.image = m3
    m3l.pack(side=tk.BOTTOM, pady=20, padx=(0,20))

    mfr.pack()

    abf = tk.Frame(abtf, background="#003459")

    des = ImageTk.PhotoImage(Image.open("./Images/des.png").resize((950, 450)))
    desl = tk.Label(abf, image=des, text="", background="#003459")
    desl.image = des
    desl.pack(side=tk.LEFT, padx=10)

    pfp = ImageTk.PhotoImage(Image.open("./Images/pfp.jpg").resize((300,450)))
    pfpl = tk.Label(abf, image=pfp, text="", background="#003459")
    pfpl.image = pfp
    pfpl.pack(side=tk.RIGHT, padx=10)

    abf.pack(pady=60, side=tk.BOTTOM)
    abtf.pack()


options_frame = tk.Frame(root, bg="#003459", borderwidth=10)
options_frame.propagate(False)
options_frame.pack(side=tk.LEFT)
options_frame.configure(width=275,height=950)

main_frame = tk.Frame(root)
main_frame.propagate(False)
main_frame.configure(width=1325,height=950, background="#15114a")
main_frame.pack()

dashboard()


sb = icon.subsample(4)
logo = tk.Label(options_frame, image=sb, bg="#003459")
logo.place(x=30, y=50)

sbn = tk.PhotoImage(file='./Images/logoname.png').subsample(5)
logoname = tk.Label(options_frame, image=sbn, bg="#003459")
logoname.place(x=35, y=250)

db = tk.PhotoImage(file='./Images/analytics_icon.png').subsample(10)
dbl = tk.Label(options_frame, image=db, bg="#003459")
dbl.place(x=5, y=385)
op1 = tk.PhotoImage(file='./Images/op1.png').subsample(2)
dashboard_btn = tk.Button(options_frame, command=lambda: dashboard(),image=op1, bg="#003459", fg="white", bd= 0, activebackground='#003459')
dashboard_btn.place(x= 50, y=400)

order = tk.PhotoImage(file='./Images/search.png').subsample(10)
orl = tk.Label(options_frame, image=order, bg="#003459")
orl.place(x=5, y=485)
op2 = tk.PhotoImage(file='./Images/op2.png').subsample(5)
orders_btn = tk.Button(options_frame, command=lambda: orders(),image=op2, bg="#003459", fg="white", bd= 0, activebackground='#003459')
orders_btn.place(x= 50, y=500)

add = tk.PhotoImage(file='./Images/add.png').subsample(10)
addl = tk.Label(options_frame, image=add, bg="#003459")
addl.place(x=5, y=585)
op3 = tk.PhotoImage(file='./Images/op3.png').subsample(5)
addorder_btn = tk.Button(options_frame, command=lambda: addorders(), image=op3, bg="#003459", fg="white", bd= 0, activebackground='#003459')
addorder_btn.place(x= 50, y=600)

modt = tk.PhotoImage(file='./Images/trash.png').subsample(10)
modl = tk.Label(options_frame, image=modt, bg="#003459")
modl.place(x=5, y=685)
op4 = tk.PhotoImage(file='./Images/op4.png').subsample(5)
modorder_btn = tk.Button(options_frame, command=lambda: modorder(), image=op4, bg="#003459", fg="white",  bd= 0, activebackground='#003459')
modorder_btn.place(x= 50, y=700)

ab = tk.PhotoImage(file='./Images/about.png').subsample(10)
abl = tk.Label(options_frame, image=ab, bg="#003459")
abl.place(x=5, y=785)
op5 = tk.PhotoImage(file='./Images/op5.png').subsample(5)
abt_btn = tk.Button(options_frame, command=lambda: abt(), image=op5, bg="#003459",  fg="white", bd=0, activebackground='#003459')
abt_btn.place(x= 50, y=800)

root.mainloop()
